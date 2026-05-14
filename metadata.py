import polars as pl
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from encoding_detect import EncodingResult, detect_encoding
from utils import Progress, check_cancel


@dataclass
class FileMetadata:
    path: Path
    size_bytes: int
    encoding: EncodingResult
    delimiter: str
    row_count: int
    column_count: int
    columns: list[str]
    dtypes: dict[str, str]
    sheet_name: Optional[str] = None


@dataclass
class SchemaDiff:
    columns_only_in_f1: list[str]
    columns_only_in_f2: list[str]
    columns_in_both: list[str]
    renamed_candidates: list[tuple[str, str, float]]   # (f1_col, f2_col, similarity)
    reorder_detected: bool
    column_order_f1: list[str]
    column_order_f2: list[str]
    compatibility_score: float  # 0-100


def load_metadata(
    path: Path,
    sheet_name: Optional[str] = None,
) -> FileMetadata:
    """
    Load file metadata: size, encoding, delimiter, row/col count, dtypes.
    CSV/TSV: auto-detects encoding and delimiter.
    Excel: requires openpyxl; sheet_name optional (defaults to first sheet).
    """
    size_bytes = path.stat().st_size
    ext = path.suffix.lower()

    if ext == ".xlsx":
        return _load_xlsx_metadata(path, sheet_name, size_bytes)

    # CSV / TSV path
    encoding_result = detect_encoding(path)
    delimiter = detect_delimiter(path, encoding_result.encoding)

    lf = pl.scan_csv(
        path,
        encoding=_polars_encoding(encoding_result.encoding),
        separator=delimiter,
        infer_schema_length=10_000,
        ignore_errors=True,
    )
    collected_schema = lf.collect_schema()
    columns = collected_schema.names()

    # Row count via lazy count (no full read)
    row_count = lf.select(pl.len()).collect().item()
    dtypes = {col: str(dtype) for col, dtype in collected_schema.items()}

    return FileMetadata(
        path=path,
        size_bytes=size_bytes,
        encoding=encoding_result,
        delimiter=delimiter,
        row_count=row_count,
        column_count=len(columns),
        columns=columns,
        dtypes=dtypes,
        sheet_name=None,
    )


def _load_xlsx_metadata(path: Path, sheet_name: Optional[str], size_bytes: int) -> FileMetadata:
    import openpyxl
    from encoding_detect import EncodingResult

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    resolved_sheet = sheet_name or wb.sheetnames[0]
    ws = wb[resolved_sheet]

    # Read header row
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    columns = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(header_row)]
    row_count = ws.max_row - 1  # subtract header

    wb.close()

    dummy_enc = EncodingResult(
        encoding="utf-8",
        confidence=1.0,
        has_bom=False,
        bom_bytes=None,
        raw_sample=b"",
    )

    return FileMetadata(
        path=path,
        size_bytes=size_bytes,
        encoding=dummy_enc,
        delimiter=",",
        row_count=max(row_count, 0),
        column_count=len(columns),
        columns=columns,
        dtypes={col: "Utf8" for col in columns},
        sheet_name=resolved_sheet,
    )


def detect_delimiter(path: Path, encoding: str, sample_lines: int = 20) -> str:
    """Detect delimiter by counting occurrences across sample lines."""
    try:
        with open(path, encoding=_safe_encoding(encoding), errors="ignore") as f:
            lines = [f.readline().rstrip("\n") for _ in range(sample_lines)]

        candidates = {",": [], "\t": [], "|": [], ";": []}
        for delim in candidates:
            counts = [line.count(delim) for line in lines if line]
            candidates[delim] = counts

        # Score: prefer delimiter with consistent non-zero count
        def score(counts: list[int]) -> float:
            if not counts or max(counts) == 0:
                return -1.0
            mean = sum(counts) / len(counts)
            variance = sum((c - mean) ** 2 for c in counts) / len(counts)
            return mean - variance  # high mean, low variance

        best = max(candidates, key=lambda d: score(candidates[d]))
        if score(candidates[best]) > 0:
            return best
        return ","
    except Exception:
        return ","


def _polars_encoding(enc: str) -> str:
    """Map detected encoding names to Polars-accepted encoding strings."""
    mapping = {
        "utf-8-sig": "utf8-lossy",
        "utf-8": "utf8",
        "utf8": "utf8",
        "latin-1": "utf8-lossy",
        "iso-8859-1": "utf8-lossy",
        "cp1252": "utf8-lossy",
        "windows-1252": "utf8-lossy",
        "utf-16-le": "utf8-lossy",
        "utf-16-be": "utf8-lossy",
    }
    return mapping.get(enc.lower(), "utf8-lossy")


def _safe_encoding(enc: str) -> str:
    """Map encoding names to Python-safe codec names."""
    mapping = {
        "utf-8-sig": "utf-8-sig",
        "latin-1": "latin-1",
        "iso-8859-1": "latin-1",
        "cp1252": "cp1252",
        "windows-1252": "cp1252",
    }
    return mapping.get(enc.lower(), enc)


def compare_schemas(m1: FileMetadata, m2: FileMetadata) -> SchemaDiff:
    """
    Compare two file schemas. Detect column additions, removals, renames, reordering.
    Returns a SchemaDiff with a compatibility score (0-100).
    """
    set1 = set(m1.columns)
    set2 = set(m2.columns)

    only_in_f1 = sorted(set1 - set2)
    only_in_f2 = sorted(set2 - set1)
    in_both = sorted(set1 & set2)

    # Rename detection via fuzzy matching on missing columns
    renamed_candidates: list[tuple[str, str, float]] = []
    for col1 in only_in_f1:
        for col2 in only_in_f2:
            max_len = max(len(col1), len(col2))
            if max_len == 0:
                continue
            sim = 1.0 - (_levenshtein(col1.lower(), col2.lower()) / max_len)
            if sim > 0.7:
                renamed_candidates.append((col1, col2, round(sim, 3)))

    # Reorder: the shared columns appear in the same relative order in both files
    shared_order_f1 = [c for c in m1.columns if c in in_both]
    shared_order_f2 = [c for c in m2.columns if c in in_both]
    reorder_detected = shared_order_f1 != shared_order_f2

    # Compatibility score (heuristic)
    score = 100.0
    score -= len(only_in_f1) * 5
    score -= len(only_in_f2) * 2
    if reorder_detected:
        score -= 5
    type_mismatches = sum(
        1 for c in in_both if m1.dtypes.get(c) != m2.dtypes.get(c)
    )
    score -= type_mismatches * 3
    score = max(0.0, min(100.0, score))

    return SchemaDiff(
        columns_only_in_f1=only_in_f1,
        columns_only_in_f2=only_in_f2,
        columns_in_both=in_both,
        renamed_candidates=renamed_candidates,
        reorder_detected=reorder_detected,
        column_order_f1=list(m1.columns),
        column_order_f2=list(m2.columns),
        compatibility_score=score,
    )


def _levenshtein(s1: str, s2: str) -> int:
    if len(s1) < len(s2):
        return _levenshtein(s2, s1)
    if len(s2) == 0:
        return len(s1)
    prev_row = list(range(len(s2) + 1))
    for i, c1 in enumerate(s1):
        curr_row = [i + 1] + [0] * len(s2)
        for j, c2 in enumerate(s2):
            curr_row[j + 1] = min(
                prev_row[j + 1] + 1,   # deletion
                curr_row[j] + 1,        # insertion
                prev_row[j] + (c1 != c2),  # substitution
            )
        prev_row = curr_row
    return prev_row[-1]


if __name__ == "__main__":
    import tempfile

    # Delimiter detection
    p_comma = Path(tempfile.mktemp(suffix=".csv"))
    p_comma.write_text("id,name,value\n1,Alice,10\n2,Bob,20\n")
    assert detect_delimiter(p_comma, "utf-8") == ",", detect_delimiter(p_comma, "utf-8")
    p_comma.unlink()

    p_tab = Path(tempfile.mktemp(suffix=".tsv"))
    p_tab.write_text("id\tname\tvalue\n1\tAlice\t10\n2\tBob\t20\n")
    assert detect_delimiter(p_tab, "utf-8") == "\t", detect_delimiter(p_tab, "utf-8")
    p_tab.unlink()

    # Schema comparison
    from dataclasses import replace
    from encoding_detect import EncodingResult
    dummy_enc = EncodingResult("utf-8", 1.0, False, None, b"")

    m1 = FileMetadata(Path("a.csv"), 0, dummy_enc, ",", 10, 3,
                      ["id", "name", "value"], {"id": "Int64", "name": "Utf8", "value": "Float64"})
    m2 = FileMetadata(Path("b.csv"), 0, dummy_enc, ",", 10, 3,
                      ["id", "name", "amount"], {"id": "Int64", "name": "Utf8", "amount": "Float64"})

    diff = compare_schemas(m1, m2)
    assert diff.columns_only_in_f1 == ["value"]
    assert diff.columns_only_in_f2 == ["amount"]
    assert "id" in diff.columns_in_both
    # "amount" ≈ "value": not expected to match, but renamed_candidates may be empty
    assert isinstance(diff.compatibility_score, float)

    print("✓ Metadata tests passed")
