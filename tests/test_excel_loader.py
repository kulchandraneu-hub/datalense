"""
Tests for excel_loader module and GET /api/sheets endpoint.

Required coverage (P4-T1 acceptance criteria):
  - list_sheets() returns correct sheet names
  - excel_to_temp_csv() produces a valid CSV with correct content
  - /api/sheets endpoint returns sheet list

Marker: quick — all fixtures are small in-memory workbooks.
"""
import csv
import sys
from pathlib import Path
import pytest
import openpyxl

sys.path.insert(0, str(Path(__file__).parent.parent))
import excel_loader  # noqa: E402

pytestmark = pytest.mark.quick


# ---------------------------------------------------------------------------
# Shared fixture: small .xlsx with two sheets
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def sample_xlsx(tmp_path_factory):
    """Two-sheet workbook: 'Employees' (3 rows) and 'Departments' (2 rows)."""
    tmp = tmp_path_factory.mktemp("excel_fixtures")
    path = tmp / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Employees"
    ws1.append(["ID", "Name", "Salary"])
    ws1.append([1, "Alice", 50000])
    ws1.append([2, "Bob", 60000])
    ws2 = wb.create_sheet("Departments")
    ws2.append(["DeptID", "DeptName"])
    ws2.append([10, "Engineering"])
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# detect_file_type
# ---------------------------------------------------------------------------

class TestDetectFileType:
    def test_csv_extension(self, tmp_path):
        p = tmp_path / "data.csv"
        p.write_text("a,b\n1,2\n")
        assert excel_loader.detect_file_type(p) == "csv"

    def test_tsv_extension(self, tmp_path):
        p = tmp_path / "data.tsv"
        p.write_text("a\tb\n1\t2\n")
        assert excel_loader.detect_file_type(p) == "tsv"

    def test_xlsx_extension_and_magic(self, sample_xlsx):
        assert excel_loader.detect_file_type(sample_xlsx) == "xlsx"

    def test_txt_falls_back_to_csv(self, tmp_path):
        p = tmp_path / "data.txt"
        p.write_text("a,b\n1,2\n")
        assert excel_loader.detect_file_type(p) == "csv"


# ---------------------------------------------------------------------------
# list_sheets  (required test #1)
# ---------------------------------------------------------------------------

class TestListSheets:
    def test_returns_correct_sheet_names(self, sample_xlsx):
        sheets = excel_loader.list_sheets(sample_xlsx)
        assert sheets == ["Employees", "Departments"]

    def test_order_preserved(self, sample_xlsx):
        sheets = excel_loader.list_sheets(sample_xlsx)
        assert sheets[0] == "Employees"
        assert sheets[1] == "Departments"

    def test_csv_returns_empty_list(self, tmp_path):
        p = tmp_path / "data.csv"
        p.write_text("id,name\n1,Alice\n")
        assert excel_loader.list_sheets(p) == []

    def test_tsv_returns_empty_list(self, tmp_path):
        p = tmp_path / "data.tsv"
        p.write_text("id\tname\n1\tAlice\n")
        assert excel_loader.list_sheets(p) == []

    def test_single_sheet_workbook(self, tmp_path):
        path = tmp_path / "one_sheet.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "OnlySheet"
        wb.save(path)
        assert excel_loader.list_sheets(path) == ["OnlySheet"]


# ---------------------------------------------------------------------------
# get_default_sheet
# ---------------------------------------------------------------------------

class TestGetDefaultSheet:
    def test_returns_first_sheet(self, sample_xlsx):
        assert excel_loader.get_default_sheet(sample_xlsx) == "Employees"

    def test_raises_for_csv(self, tmp_path):
        p = tmp_path / "data.csv"
        p.write_text("id,name\n1,Alice\n")
        with pytest.raises(ValueError):
            excel_loader.get_default_sheet(p)


# ---------------------------------------------------------------------------
# excel_to_temp_csv  (required test #2)
# ---------------------------------------------------------------------------

class TestExcelToTempCsv:
    def test_produces_valid_csv(self, sample_xlsx, tmp_path):
        tmp = excel_loader.excel_to_temp_csv(sample_xlsx, "Employees", tmp_path)
        try:
            assert tmp.exists()
            assert tmp.suffix == ".csv"
            content = tmp.read_text(encoding="utf-8")
            assert "ID" in content
            assert "Alice" in content
            assert "50000" in content
        finally:
            tmp.unlink(missing_ok=True)

    def test_correct_row_count_and_headers(self, sample_xlsx, tmp_path):
        tmp = excel_loader.excel_to_temp_csv(sample_xlsx, "Employees", tmp_path)
        try:
            with open(tmp, encoding="utf-8", newline="") as f:
                rows = list(csv.reader(f))
            assert len(rows) == 3                           # header + 2 data rows
            assert rows[0] == ["ID", "Name", "Salary"]
            assert rows[1][1] == "Alice"
            assert rows[2][1] == "Bob"
        finally:
            tmp.unlink(missing_ok=True)

    def test_second_sheet_content(self, sample_xlsx, tmp_path):
        tmp = excel_loader.excel_to_temp_csv(sample_xlsx, "Departments", tmp_path)
        try:
            with open(tmp, encoding="utf-8", newline="") as f:
                rows = list(csv.reader(f))
            assert rows[0] == ["DeptID", "DeptName"]
            assert rows[1][1] == "Engineering"
        finally:
            tmp.unlink(missing_ok=True)

    def test_temp_file_written_to_requested_dir(self, sample_xlsx, tmp_path):
        tmp = excel_loader.excel_to_temp_csv(sample_xlsx, "Employees", tmp_path)
        try:
            assert tmp.parent == tmp_path
        finally:
            tmp.unlink(missing_ok=True)

    def test_non_excel_raises_value_error(self, tmp_path):
        p = tmp_path / "data.csv"
        p.write_text("id,name\n1,Alice\n")
        with pytest.raises(ValueError, match="Not an Excel file"):
            excel_loader.excel_to_temp_csv(p, "Sheet1", tmp_path)

    def test_none_values_become_empty_string(self, tmp_path):
        path = tmp_path / "with_nulls.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["ID", "Value"])
        ws.append([1, None])
        wb.save(path)
        tmp = excel_loader.excel_to_temp_csv(path, "Data", tmp_path)
        try:
            with open(tmp, encoding="utf-8", newline="") as f:
                rows = list(csv.reader(f))
            assert rows[1] == ["1", ""]   # None → empty string
        finally:
            tmp.unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# GET /api/sheets endpoint  (required test #3)
# ---------------------------------------------------------------------------

def _make_client():
    """Return a TestClient for the FastAPI app, or None if unavailable."""
    try:
        from fastapi.testclient import TestClient
        from web.api import app
        return TestClient(app)
    except Exception:
        return None


_client = _make_client()
_skip_api = pytest.mark.skipif(
    _client is None,
    reason="FastAPI TestClient not available (httpx or requests missing)",
)


@_skip_api
class TestApiSheetsEndpoint:
    """GET /api/sheets — returns sheet names or [] for CSV/TSV."""

    def test_returns_sheet_names_for_xlsx(self, sample_xlsx):
        resp = _client.get("/api/sheets", params={"path": str(sample_xlsx)})
        assert resp.status_code == 200
        data = resp.json()
        assert "sheets" in data
        assert data["sheets"] == ["Employees", "Departments"]

    def test_returns_empty_list_for_csv(self, tmp_path):
        p = tmp_path / "test.csv"
        p.write_text("id,name\n1,Alice\n")
        resp = _client.get("/api/sheets", params={"path": str(p)})
        assert resp.status_code == 200
        assert resp.json()["sheets"] == []

    def test_404_for_missing_file(self):
        resp = _client.get("/api/sheets", params={"path": "C:/no/such/file.xlsx"})
        assert resp.status_code == 404
