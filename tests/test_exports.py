"""
Regression tests for export output structure.

Coverage:
  - Excel: {col}_before / {col}_after interleaved headers (P1-T4 regression)
  - Excel: change_type and key column present; bare column names NOT present
  - Excel: row count matches sample_diffs length
  - JSON: top-level keys, summary counts match DiffResult, column_diffs structure
  - CSV: header includes change_type, key columns, {col}_f1 / {col}_f2 pairs

Marker: quick — runs against clean_result (5-row synthetic fixture).
"""
import csv
import io
import sys
import pytest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))
from reporters import render_excel_diff, render_json_diff, render_csv_diff  # noqa: E402

pytestmark = pytest.mark.quick


class TestExcelExport:
    """Excel diff export: P1-T4 regression — before/after column pairs, correct layout."""

    def test_file_is_created(self, clean_result, tmp_path):
        path = tmp_path / "diff.xlsx"
        render_excel_diff(clean_result.diff, path)
        assert path.exists()
        assert path.stat().st_size > 0

    def test_before_after_headers_present(self, clean_result, tmp_path):
        """Each data column must produce {col}_before and {col}_after headers."""
        path = tmp_path / "diff.xlsx"
        render_excel_diff(clean_result.diff, path)
        import openpyxl
        headers = [c.value for c in openpyxl.load_workbook(path).active[1]]
        for col in clean_result.diff.column_diffs:
            assert f"{col}_before" in headers, f"Missing {col}_before in Excel headers"
            assert f"{col}_after" in headers, f"Missing {col}_after in Excel headers"

    def test_change_type_column_present(self, clean_result, tmp_path):
        path = tmp_path / "diff.xlsx"
        render_excel_diff(clean_result.diff, path)
        import openpyxl
        headers = [c.value for c in openpyxl.load_workbook(path).active[1]]
        assert "change_type" in headers

    def test_key_column_in_headers(self, clean_result, tmp_path):
        path = tmp_path / "diff.xlsx"
        render_excel_diff(clean_result.diff, path)
        import openpyxl
        headers = [c.value for c in openpyxl.load_workbook(path).active[1]]
        for key_col in clean_result.diff.key_columns:
            assert key_col in headers

    def test_no_bare_column_headers(self, clean_result, tmp_path):
        """
        Pre-P1-T4 regression: headers must not contain bare column names
        (e.g., 'Value') — only {col}_before / {col}_after variants are allowed.
        """
        path = tmp_path / "diff.xlsx"
        render_excel_diff(clean_result.diff, path)
        import openpyxl
        headers = [c.value for c in openpyxl.load_workbook(path).active[1]]
        data_cols = list(clean_result.diff.column_diffs.keys())
        for col in data_cols:
            assert col not in headers, (
                f"Bare header '{col}' found in Excel; expected only {col}_before/{col}_after"
            )

    def test_data_row_count_matches_sample(self, clean_result, tmp_path):
        path = tmp_path / "diff.xlsx"
        render_excel_diff(clean_result.diff, path)
        import openpyxl
        ws = openpyxl.load_workbook(path).active
        data_rows = ws.max_row - 1  # subtract header row
        assert data_rows == len(clean_result.diff.sample_diffs)


class TestJsonExport:
    """JSON export structure and count consistency with DiffResult."""

    def test_top_level_keys(self, clean_result):
        data = render_json_diff(clean_result.diff)
        assert "summary" in data
        assert "column_diffs" in data
        assert "sample_rows" in data

    def test_summary_counts_match_diff(self, clean_result):
        data = render_json_diff(clean_result.diff)
        s = data["summary"]
        assert s["added"] == clean_result.diff.added_rows
        assert s["removed"] == clean_result.diff.removed_rows
        assert s["modified"] == clean_result.diff.modified_rows
        assert s["formatting_only"] == clean_result.diff.formatting_only_rows

    def test_summary_key_columns(self, clean_result):
        data = render_json_diff(clean_result.diff)
        assert data["summary"]["key_columns"] == clean_result.diff.key_columns

    def test_column_diffs_required_keys(self, clean_result):
        data = render_json_diff(clean_result.diff)
        for col_name, stats in data["column_diffs"].items():
            for key in ("modified", "formatting_only", "null_introduced", "null_resolved", "change_rate"):
                assert key in stats, f"Key '{key}' missing in column_diffs['{col_name}']"

    def test_change_rate_is_float(self, clean_result):
        data = render_json_diff(clean_result.diff)
        for col_name, stats in data["column_diffs"].items():
            assert isinstance(stats["change_rate"], float), (
                f"change_rate for '{col_name}' is not a float"
            )


class TestCsvExport:
    """CSV export header format and column presence."""

    def test_csv_has_header_row(self, clean_result):
        content = render_csv_diff(clean_result.diff)
        rows = list(csv.reader(io.StringIO(content)))
        assert len(rows) >= 1

    def test_change_type_in_header(self, clean_result):
        content = render_csv_diff(clean_result.diff)
        header = next(csv.reader(io.StringIO(content)))
        assert "change_type" in header

    def test_key_columns_in_header(self, clean_result):
        content = render_csv_diff(clean_result.diff)
        header = next(csv.reader(io.StringIO(content)))
        for key_col in clean_result.diff.key_columns:
            assert key_col in header

    def test_f1_f2_column_pairs_in_header(self, clean_result):
        content = render_csv_diff(clean_result.diff)
        header = next(csv.reader(io.StringIO(content)))
        for col in clean_result.diff.column_diffs:
            assert f"{col}_f1" in header, f"Missing {col}_f1 in CSV header"
            assert f"{col}_f2" in header, f"Missing {col}_f2 in CSV header"
