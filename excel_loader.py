"""
Excel file support for DataLens.

Provides file-type detection, sheet enumeration, and Excel-to-CSV conversion.
The core comparison engine (differ.py, profiler.py, compare.py) is unchanged;
Excel files are converted to a temporary CSV before entering the pipeline.
"""

import csv
import os
import tempfile
from pathlib import Path


def detect_file_type(path: Path) -> str:
    """
    Detect file type from extension and magic bytes.
    Returns 'csv', 'tsv', 'xlsx', or 'xls'.
    Extension is trusted for CSV/TSV; Excel files are also validated by magic bytes.
    """
    ext = path.suffix.lower()

    if ext == ".xlsx":
        try:
            with open(path, "rb") as fh:
                magic = fh.read(4)
            # .xlsx is a ZIP archive; all ZIPs start with PK signature bytes
            if magic[:2] == b"PK":
                return "xlsx"
        except OSError:
            pass
        return "xlsx"  # trust extension if file is unreadable

    if ext == ".xls":
        try:
            with open(path, "rb") as fh:
                magic = fh.read(4)
            # OLE2 compound document (legacy Excel .xls)
            if magic == b"\xd0\xcf\x11\xe0":
                return "xls"
        except OSError:
            pass
        return "xls"

    if ext == ".tsv":
        return "tsv"

    return "csv"


def list_sheets(path: Path) -> list[str]:
    """
    Return sheet names for an Excel file.
    Returns [] for CSV / TSV files.
    Uses openpyxl read_only=True — never loads the full workbook into RAM.
    """
    ft = detect_file_type(path)

    if ft == "xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            names = list(wb.sheetnames)
            wb.close()
            return names
        except Exception:
            return []

    if ft == "xls":
        # xlrd is optional — not in default requirements
        try:
            import xlrd
            wb = xlrd.open_workbook(str(path))
            return list(wb.sheet_names())
        except Exception:
            return []

    return []


def get_default_sheet(path: Path) -> str:
    """Return the name of the first sheet. Raises ValueError if no sheets are found."""
    sheets = list_sheets(path)
    if not sheets:
        raise ValueError(f"No sheets found in {path}")
    return sheets[0]


def excel_to_temp_csv(path: Path, sheet_name: str, temp_dir: Path) -> Path:
    """
    Convert one Excel sheet to a temp CSV file written into temp_dir.
    Returns the Path to the temp file. Caller must delete it (use try/finally).

    INV-3: mkstemp() + os.close() before any write for Windows handle safety.
    """
    ft = detect_file_type(path)
    if ft not in ("xlsx", "xls"):
        raise ValueError(f"Not an Excel file: {path}")

    # INV-3: release the OS handle before opening for write on Windows
    fd, tmp_str = tempfile.mkstemp(dir=str(temp_dir), suffix=".csv")
    os.close(fd)
    tmp_path = Path(tmp_str)

    try:
        if ft == "xlsx":
            _xlsx_to_csv(path, sheet_name, tmp_path)
        else:
            _xls_to_csv(path, sheet_name, tmp_path)
    except Exception:
        tmp_path.unlink(missing_ok=True)
        raise

    return tmp_path


def _xlsx_to_csv(xlsx_path: Path, sheet_name: str, out_path: Path) -> None:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    with open(out_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([str(v) if v is not None else "" for v in row])
    wb.close()


def _xls_to_csv(xls_path: Path, sheet_name: str, out_path: Path) -> None:
    import xlrd  # optional dependency
    wb = xlrd.open_workbook(str(xls_path))
    ws = wb.sheet_by_name(sheet_name)
    with open(out_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        for row_idx in range(ws.nrows):
            writer.writerow([
                str(ws.cell_value(row_idx, col_idx))
                for col_idx in range(ws.ncols)
            ])
