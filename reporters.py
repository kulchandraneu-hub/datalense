from pathlib import Path
from typing import Optional
import json

from differ import DiffResult, RowDiff
from validator import ValidationReport


def render_html_report(
    validation_f1: ValidationReport,
    validation_f2: Optional[ValidationReport],
    diff: Optional[DiffResult],
) -> str:
    """Generate a self-contained dark-theme HTML report (no external dependencies)."""

    def _sev_class(sev: str) -> str:
        return {"INFO": "info", "WARNING": "warn", "ERROR": "error", "CRITICAL": "critical"}.get(sev, "info")

    def _checks_html(report: ValidationReport) -> str:
        parts = []
        for chk in report.checks:
            cnt = f" ({chk.affected_count:,} rows)" if chk.affected_count is not None else ""
            icon = {"INFO": "ℹ", "WARNING": "⚠", "ERROR": "✖", "CRITICAL": "☠"}.get(chk.severity, "•")
            parts.append(
                f'<div class="check {_sev_class(chk.severity)}">'
                f'{icon} <strong>{chk.name}</strong> [{chk.severity}]: {chk.message}{cnt}'
                f'</div>'
            )
        return "\n".join(parts) if parts else "<p>No issues found.</p>"

    f1_path = validation_f1.profile.metadata.path
    f1_rows = f"{validation_f1.total_count:,}"
    f1_cols = validation_f1.column_count
    f1_score = (
        f"{validation_f1.compatibility_score:.0f}/100"
        if validation_f1.compatibility_score is not None else "N/A"
    )

    diff_section = ""
    if diff:
        col_rows = "\n".join(
            f"<tr><td>{n}</td><td>{s.modified_count:,}</td>"
            f"<td>{s.formatting_only_count:,}</td>"
            f"<td>{s.change_rate * 100:.1f}%</td></tr>"
            for n, s in diff.column_diffs.items()
        )
        diff_section = f"""
        <h2>Row-Level Changes</h2>
        <div class="stat-row">
          <div class="stat green">Added<br><span>{diff.added_rows:,}</span></div>
          <div class="stat red">Removed<br><span>{diff.removed_rows:,}</span></div>
          <div class="stat yellow">Modified<br><span>{diff.modified_rows:,}</span></div>
          <div class="stat grey">Fmt Only<br><span>{diff.formatting_only_rows:,}</span></div>
        </div>
        <h3>Column Changes</h3>
        <table>
          <tr><th>Column</th><th>Modified</th><th>Fmt-only</th><th>Change rate</th></tr>
          {col_rows}
        </table>
        """

    f2_section = ""
    if validation_f2:
        f2_path = validation_f2.profile.metadata.path
        f2_section = f"""
        <h2>File 2: {f2_path.name}</h2>
        <p>Rows: {validation_f2.total_count:,} | Columns: {validation_f2.column_count}</p>
        <h3>Validation Checks</h3>
        {_checks_html(validation_f2)}
        """

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <title>Data Quality Report</title>
  <style>
    body {{ background:#0F172A; color:#F1F5F9; font-family:monospace; padding:24px; max-width:960px; margin:auto }}
    h1 {{ color:#3B82F6; border-bottom:1px solid #1E293B; padding-bottom:8px }}
    h2 {{ color:#7DD3FC; margin-top:24px }}
    h3 {{ color:#94A3B8 }}
    .check {{ padding:8px 12px; margin:4px 0; border-left:4px solid; border-radius:2px }}
    .check.info     {{ border-color:#3B82F6; background:#1E3A5F }}
    .check.warn     {{ border-color:#F59E0B; background:#4A3A0E }}
    .check.error    {{ border-color:#EF4444; background:#4A1515 }}
    .check.critical {{ border-color:#8B5CF6; background:#2D1B4E }}
    table {{ border-collapse:collapse; width:100%; margin:12px 0 }}
    th,td {{ border:1px solid #334155; padding:8px; text-align:left }}
    th {{ background:#1E293B }}
    .stat-row {{ display:flex; gap:12px; margin:12px 0 }}
    .stat {{ flex:1; padding:12px; border-radius:4px; text-align:center; font-size:0.85em }}
    .stat span {{ font-size:1.6em; font-weight:bold; display:block; margin-top:4px }}
    .stat.green  {{ background:#14532D }}
    .stat.red    {{ background:#7F1D1D }}
    .stat.yellow {{ background:#78350F }}
    .stat.grey   {{ background:#1F2937 }}
  </style>
</head>
<body>
  <h1>Data Quality Report</h1>
  <h2>File 1: {f1_path.name}</h2>
  <p>Rows: {f1_rows} | Columns: {f1_cols} | Compatibility: {f1_score}</p>
  <h3>Validation Checks</h3>
  {_checks_html(validation_f1)}
  {f2_section}
  {diff_section}
</body>
</html>"""


def render_excel_diff(diff: DiffResult, output_path: Path) -> None:
    """Write diff to Excel. Each data column gets {col}_before and {col}_after columns.
    Changed cells within modified rows are highlighted individually."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel export: pip install openpyxl")

    from datetime import datetime

    def _auto_width(ws_target) -> None:
        for i, col_cells in enumerate(ws_target.columns, 1):
            max_len = max((len(str(cell.value or "")) for cell in col_cells), default=0)
            ws_target.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)

    wb = openpyxl.Workbook()

    # --- Summary sheet (tab index 0 — first tab in Excel) ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["Field", "Value"])
    for cell in ws_sum[1]:
        cell.font = Font(bold=True)
    for label, value in [
        ("File 1 rows",          diff.total_rows_f1),
        ("File 2 rows",          diff.total_rows_f2),
        ("Added rows",           diff.added_rows),
        ("Removed rows",         diff.removed_rows),
        ("Modified rows",        diff.modified_rows),
        ("Formatting-only rows", diff.formatting_only_rows),
        ("Compare date/time",    datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Key columns",          ", ".join(diff.key_columns)),
    ]:
        ws_sum.append([label, value])
    _auto_width(ws_sum)

    # --- Diff sheet (tab index 1; set as active so tests using .active find it) ---
    ws = wb.create_sheet("Diff")
    wb._active_sheet_index = 1  # keep Summary first but make Diff the active sheet

    row_fills = {
        "added":           PatternFill("solid", fgColor="D4EDDA"),
        "removed":         PatternFill("solid", fgColor="F8D7DA"),
        "modified":        PatternFill("solid", fgColor="FFF3CD"),
        "formatting_only": PatternFill("solid", fgColor="F8F9FA"),
    }
    changed_cell_fill = PatternFill("solid", fgColor="F59E0B")

    data_cols = list(diff.column_diffs.keys())
    key_count = len(diff.key_columns)

    # Header: [key..., change_type, col1_before, col1_after, col2_before, col2_after, ...]
    header = [*diff.key_columns, "change_type"]
    for col in data_cols:
        header.append(f"{col}_before")
        header.append(f"{col}_after")
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Freeze header row so column names stay visible when scrolling
    ws.freeze_panes = "A2"

    # 0-based index of each column's _after cell in the row tuple.
    # Layout: 0..key_count-1 = keys, key_count = change_type,
    #         key_count+1+2i = col_i _before, key_count+2+2i = col_i _after
    col_after_idx = {col: key_count + 2 + 2 * i for i, col in enumerate(data_cols)}

    for row in diff.sample_diffs:
        vals = [row.key_value, row.change_type]
        for col in data_cols:
            vals.append(row.f1_values.get(col, ""))
            vals.append(row.f2_values.get(col, ""))
        ws.append(vals)

        excel_row = ws[ws.max_row]
        fill = row_fills.get(row.change_type)
        if fill:
            for cell in excel_row:
                cell.fill = fill

        # Highlight the _after cell for each semantically-changed column
        if row.change_type == "modified" and row.columns_changed:
            for col in row.columns_changed:
                after_i = col_after_idx.get(col)
                if after_i is not None:
                    excel_row[after_i].fill = changed_cell_fill

    _auto_width(ws)

    wb.save(output_path)


def render_json_diff(diff: DiffResult) -> dict:
    """Return diff as a JSON-serialisable dict."""
    return {
        "summary": {
            "added": diff.added_rows,
            "removed": diff.removed_rows,
            "modified": diff.modified_rows,
            "formatting_only": diff.formatting_only_rows,
            "total_f1": diff.total_rows_f1,
            "total_f2": diff.total_rows_f2,
            "confidence_score": diff.confidence_score,
            "key_columns": diff.key_columns,
        },
        "column_diffs": {
            name: {
                "modified": s.modified_count,
                "formatting_only": s.formatting_only_count,
                "null_introduced": s.null_introduced_count,
                "null_resolved": s.null_resolved_count,
                "change_rate": round(s.change_rate, 4),
            }
            for name, s in diff.column_diffs.items()
        },
        "sample_rows": [
            {
                "key": r.key_value,
                "change_type": r.change_type,
                "columns_changed": r.columns_changed,
                "f1": r.f1_values,
                "f2": r.f2_values,
            }
            for r in diff.sample_diffs[:100]
        ],
    }


def render_csv_diff(
    diff: DiffResult,
    diff_lf: Optional["pl.LazyFrame"] = None,
    output_path: Optional[Path] = None,
) -> str:
    """Return diff sample rows as CSV string, or stream full diff to disk via sink_csv.

    If diff_lf and output_path are both provided the full diff LazyFrame is written
    directly to disk with sink_csv (no intermediate RAM allocation). The caller is
    responsible for shaping diff_lf into the desired column layout before passing it.
    Returns an empty string in that case — the file content is on disk at output_path.
    If diff_lf/output_path are absent, falls back to the sample-based StringIO path.
    """
    if diff_lf is not None and output_path is not None:
        import polars as _pl
        diff_lf.sink_csv(output_path)
        return ""

    import csv, io

    buf = io.StringIO()
    writer = csv.writer(buf)

    col_names = list(diff.column_diffs.keys())
    writer.writerow([*diff.key_columns, "change_type", *[f"{c}_f1" for c in col_names], *[f"{c}_f2" for c in col_names]])

    for row in diff.sample_diffs:
        f1_vals = [row.f1_values.get(c, "") for c in col_names]
        f2_vals = [row.f2_values.get(c, "") for c in col_names]
        writer.writerow([row.key_value, row.change_type, *f1_vals, *f2_vals])

    return buf.getvalue()


if __name__ == "__main__":
    print("OK Reporters module ready for integration")
